import os
import random
from typing import Union
from typing import List, Dict
from openpyxl import load_workbook


class AccountManager:
    def __init__(self, accounts_file: str):
        self.accounts_file = accounts_file
        self.accounts = self._load_accounts()

    def _load_accounts(self) -> List[Dict]:
        accounts = []
        try:
            wb = load_workbook(filename=self.accounts_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
                if len(row) >= 3:
                    account_name, evm_wallet, proxy = row[:3]
                    if account_name and evm_wallet and proxy:
                        accounts.append({
                            'account_name': account_name,
                            'evm_wallet': evm_wallet,
                            'proxy': proxy,
                        })
                    else:
                        pass
            wb.close()
        except FileNotFoundError:
            print(f"Accounts file not found: {self.accounts_file}")
        except Exception as e:
            print(f"Error loading accounts: {e}")
        return accounts


    def get_accounts(self, range_list: Union[List[Union[int, str]], int] = None, shuffle: bool = False) -> List[Dict]:
        if range_list is None or len(range_list) == 0:
            accounts = self.accounts
        elif isinstance(range_list, int):
            accounts = [self.accounts[range_list - 1]] if 0 < range_list <= len(self.accounts) else []
        else:
            accounts = []
            for item in range_list:
                if isinstance(item, int):
                    if 0 < item <= len(self.accounts):
                        accounts.append(self.accounts[item - 1])
                elif isinstance(item, str) and '-' in item:
                    start, end = map(int, item.split('-'))
                    start = max(1, start)
                    end = min(len(self.accounts), end)
                    accounts.extend(self.accounts[start-1:end])
                else:
                    print(f"Invalid range specification: {item}")

        if shuffle:
            random.shuffle(accounts)

        return accounts


    def __len__(self):
        return len(self.accounts)

    def __getitem__(self, index):
        return self.accounts[index]